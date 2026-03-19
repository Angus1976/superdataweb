"""InterviewEntityExtractor — wraps existing AI extraction capabilities.

Provides document-level and message-level entity extraction, plus
merge logic for combining multiple extraction results.
"""

from __future__ import annotations

import io
import json
from typing import Any

from src.interview.models import (
    Entity,
    ExtractionResult,
    Relation,
    Rule,
)

# Supported file types for document extraction
SUPPORTED_FILE_TYPES = {"docx", "xlsx", "pdf"}


class InterviewEntityExtractor:
    """封装现有 src/ai/ 模块的实体提取能力。"""

    def __init__(self, ai_extractor: Any = None) -> None:
        """Inject an existing AI extractor instance (from src/ai/)."""
        self._ai = ai_extractor

    async def extract_from_document(
        self, file_content: bytes, file_type: str
    ) -> ExtractionResult:
        """从上传文档中提取实体。

        Parses the document using the appropriate library, then delegates
        to the AI extractor for entity/rule/relation extraction.
        """
        if file_type not in SUPPORTED_FILE_TYPES:
            raise ValueError(
                f"Unsupported file type: {file_type}. "
                f"Supported: {', '.join(sorted(SUPPORTED_FILE_TYPES))}"
            )

        text = self._parse_document(file_content, file_type)
        return await self._extract_from_text(text)

    async def extract_from_message(self, message: str) -> ExtractionResult:
        """从对话消息中提取实体（供 intelligent-interview 使用）。"""
        return await self._extract_from_text(message)

    async def merge_extractions(
        self, results: list[ExtractionResult]
    ) -> ExtractionResult:
        """合并多次提取结果，基于 id 去重。"""
        seen_entities: dict[str, Entity] = {}
        seen_rules: dict[str, Rule] = {}
        seen_relations: dict[str, Relation] = {}

        for r in results:
            for e in r.entities:
                seen_entities.setdefault(e.id, e)
            for rule in r.rules:
                seen_rules.setdefault(rule.id, rule)
            for rel in r.relations:
                seen_relations.setdefault(rel.id, rel)

        total_conf = sum(r.confidence for r in results) if results else 0.0
        avg_conf = total_conf / len(results) if results else 0.0

        return ExtractionResult(
            entities=list(seen_entities.values()),
            rules=list(seen_rules.values()),
            relations=list(seen_relations.values()),
            confidence=round(avg_conf, 4),
        )

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _parse_document(self, content: bytes, file_type: str) -> str:
        """Parse document bytes into plain text."""
        if file_type == "docx":
            return self._parse_docx(content)
        elif file_type == "xlsx":
            return self._parse_xlsx(content)
        elif file_type == "pdf":
            return self._parse_pdf(content)
        raise ValueError(f"Unknown file type: {file_type}")

    @staticmethod
    def _parse_docx(content: bytes) -> str:
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            raise RuntimeError("python-docx is required for .docx parsing")
        except Exception as exc:
            raise ValueError(f"Failed to parse docx: {exc}") from exc

    @staticmethod
    def _parse_xlsx(content: bytes) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        lines.append(" | ".join(cells))
            return "\n".join(lines)
        except ImportError:
            raise RuntimeError("openpyxl is required for .xlsx parsing")
        except Exception as exc:
            raise ValueError(f"Failed to parse xlsx: {exc}") from exc

    @staticmethod
    def _parse_pdf(content: bytes) -> str:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(content))
            return "\n".join(
                page.extract_text() or "" for page in reader.pages
            )
        except ImportError:
            raise RuntimeError("PyPDF2 is required for .pdf parsing")
        except Exception as exc:
            raise ValueError(f"Failed to parse pdf: {exc}") from exc

    async def _extract_from_text(self, text: str) -> ExtractionResult:
        """Delegate to AI extractor or return a stub result."""
        if self._ai is not None:
            return await self._ai.extract(text)
        # Stub: return empty result when no AI extractor is injected
        return ExtractionResult(confidence=0.0)
