"""OfflineImporter — import offline interview data and merge with online results."""

from __future__ import annotations

import io
import json
from typing import Any

from pydantic import BaseModel

from src.interview.models import ExtractionResult, ValidationResult


class ImportResult(BaseModel):
    """离线文件解析结果。"""
    entities: list[dict[str, Any]] = []
    rules: list[dict[str, Any]] = []
    relations: list[dict[str, Any]] = []
    source_file: str
    row_count: int = 0


class MergedData(BaseModel):
    """合并后的数据。"""
    entities: list[dict[str, Any]] = []
    rules: list[dict[str, Any]] = []
    relations: list[dict[str, Any]] = []
    online_count: int = 0
    offline_count: int = 0


class ImportValidationError(BaseModel):
    """导入校验错误。"""
    row: int
    field: str
    reason: str


SUPPORTED_IMPORT_TYPES = {"xlsx", "json"}


class OfflineImporter:
    """导入离线访谈数据并与在线结果合并。"""

    def validate_file(self, file_name: str, file_type: str) -> ValidationResult:
        """Validate file format."""
        if file_type not in SUPPORTED_IMPORT_TYPES:
            return ValidationResult(
                is_valid=False,
                errors=[f"Unsupported format: .{file_type}. Supported: {', '.join(sorted(SUPPORTED_IMPORT_TYPES))}"],
            )
        return ValidationResult(is_valid=True)

    async def import_file(
        self, file_content: bytes, file_type: str, file_name: str
    ) -> ImportResult:
        """Parse Excel or JSON file into standardized ImportResult."""
        if file_type == "json":
            return self._parse_json(file_content, file_name)
        elif file_type == "xlsx":
            return self._parse_xlsx(file_content, file_name)
        raise ValueError(f"Unsupported file type: {file_type}")

    async def merge_with_online(
        self,
        project_id: str,
        import_result: ImportResult,
        online_results: list[ExtractionResult] | None = None,
    ) -> MergedData:
        """Merge offline data with online extraction results (deduplicate by id)."""
        seen_entities: dict[str, dict] = {}
        seen_rules: dict[str, dict] = {}
        seen_relations: dict[str, dict] = {}

        online_count = 0
        if online_results:
            for r in online_results:
                for e in r.entities:
                    d = e.model_dump()
                    seen_entities.setdefault(d["id"], d)
                    online_count += 1
                for rule in r.rules:
                    d = rule.model_dump()
                    seen_rules.setdefault(d["id"], d)
                for rel in r.relations:
                    d = rel.model_dump()
                    seen_relations.setdefault(d["id"], d)

        offline_count = 0
        for e in import_result.entities:
            seen_entities.setdefault(e.get("id", ""), e)
            offline_count += 1
        for r in import_result.rules:
            seen_rules.setdefault(r.get("id", ""), r)
        for rel in import_result.relations:
            seen_relations.setdefault(rel.get("id", ""), rel)

        return MergedData(
            entities=list(seen_entities.values()),
            rules=list(seen_rules.values()),
            relations=list(seen_relations.values()),
            online_count=online_count,
            offline_count=offline_count,
        )

    @staticmethod
    def _parse_json(content: bytes, file_name: str) -> ImportResult:
        try:
            data = json.loads(content)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid JSON: {exc}") from exc

        if not isinstance(data, dict):
            raise ValueError("JSON root must be an object")

        return ImportResult(
            entities=data.get("entities", []),
            rules=data.get("rules", []),
            relations=data.get("relations", []),
            source_file=file_name,
            row_count=len(data.get("entities", [])) + len(data.get("rules", [])),
        )

    @staticmethod
    def _parse_xlsx(content: bytes, file_name: str) -> ImportResult:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
        except ImportError:
            raise RuntimeError("openpyxl is required for .xlsx parsing")
        except Exception as exc:
            raise ValueError(f"Failed to parse xlsx: {exc}") from exc

        entities: list[dict] = []
        ws = wb.active
        if ws:
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h).lower() if h else "" for h in rows[0]]
                for i, row in enumerate(rows[1:], start=2):
                    record = dict(zip(headers, [str(c) if c else "" for c in row]))
                    if record.get("name"):
                        entities.append({
                            "id": record.get("id", f"entity_{i}"),
                            "name": record["name"],
                            "type": record.get("type", "unknown"),
                        })

        return ImportResult(
            entities=entities,
            rules=[],
            relations=[],
            source_file=file_name,
            row_count=len(entities),
        )
